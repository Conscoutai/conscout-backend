from __future__ import annotations

import hashlib
import io
import logging
import os
import re
from collections import defaultdict
from datetime import date, datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable, Optional
from uuid import uuid4

from fastapi import HTTPException

from core.auth_context import AuthenticatedUser
from core.config import site_budget_dir
from core.database import (
    budget_audit_events_collection,
    budget_boq_items_collection,
    budget_boqs_collection,
    budget_invoices_collection,
    budget_variations_collection,
    budget_verification_runs_collection,
    material_documents_collection,
    project_materials_collection,
)
from services.progress.materials.material_service import (
    _extract_header as extract_material_header,
    _extract_pdf_pages as extract_pdf_pages,
    build_material_ledger,
    extract_structured_lines,
    get_material_summary,
    normalize_description,
    normalize_unit,
)
from services.progress.work_schedule.analytics_service import (
    build_baseline_comparison,
)
from services.progress.work_schedule.baseline_service import (
    project_currency_code,
    resolve_project,
)


MAX_FILE_BYTES = 35 * 1024 * 1024
ALLOWED_EXTENSIONS = {".pdf", ".xlsx", ".xlsm"}
INVOICE_ACTIVE_STATUSES = {
    "needs_review",
    "reviewed",
    "verified",
    "certified",
    "paid",
    "on_hold",
    "correction_requested",
}
INVOICE_HISTORY_STATUSES = {"certified", "paid"}
DECISION_ACTIONS = {"certify", "hold", "request_correction", "reject"}
VERIFICATION_TOLERANCE = 0.01

logger = logging.getLogger(__name__)


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    try:
        cleaned = str(value).replace(",", "").replace("%", "").strip()
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = f"-{cleaned[1:-1]}"
        return float(cleaned)
    except (TypeError, ValueError):
        return 0.0


def rounded(value: Any, digits: int = 4) -> float:
    return round(number(value), digits)


def revision_from_filename(filename: Any) -> str:
    matches = re.findall(
        r"\bREV(?:ISION)?[._ -]*[A-Z0-9]+",
        Path(str(filename or "")).stem,
        re.IGNORECASE,
    )
    if not matches:
        return ""
    return matches[-1].strip().upper().replace(".", "-")


def parse_iso_date(value: Any, *, field: str = "date", required: bool = False) -> str:
    raw = str(value or "").strip()
    if not raw:
        if required:
            raise HTTPException(422, f"{field} is required")
        return ""
    try:
        return date.fromisoformat(raw[:10]).isoformat()
    except ValueError as error:
        raise HTTPException(422, f"{field} must use YYYY-MM-DD") from error


def public_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, dict):
        return {
            str(key): public_value(item)
            for key, item in value.items()
            if str(key) not in {"_id", "storage_path"}
        }
    if isinstance(value, (list, tuple)):
        return [public_value(item) for item in value]
    return value


def _public(document: Optional[dict[str, Any]]) -> dict[str, Any]:
    return public_value(document or {})


def _audit(
    *,
    project_id: str,
    entity_type: str,
    entity_id: str,
    event_type: str,
    user: AuthenticatedUser,
    details: Optional[dict[str, Any]] = None,
) -> None:
    budget_audit_events_collection.insert_one(
        {
            "event_id": f"budget_audit_{uuid4().hex}",
            "project_id": project_id,
            "entity_type": entity_type,
            "entity_id": entity_id,
            "event_type": event_type,
            "actor_user_id": user.user_id,
            "actor_email": user.email,
            "actor_name": user.name,
            "details": details or {},
            "created_at": utc_now(),
        }
    )


def _safe_filename(filename: str, fallback: str) -> str:
    candidate = Path(filename or fallback).name
    return candidate or fallback


def _validate_upload(filename: str, raw_bytes: bytes) -> tuple[str, str]:
    safe_name = _safe_filename(filename, "budget-document.pdf")
    extension = Path(safe_name).suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, "Budget documents must be PDF or Excel (.xlsx/.xlsm)")
    if not raw_bytes:
        raise HTTPException(400, "The uploaded file is empty")
    if len(raw_bytes) > MAX_FILE_BYTES:
        raise HTTPException(413, "Budget documents are limited to 35 MB")
    if extension == ".pdf" and not raw_bytes.startswith(b"%PDF-"):
        raise HTTPException(400, "The uploaded file is not a valid PDF")
    if extension in {".xlsx", ".xlsm"} and not raw_bytes.startswith(b"PK"):
        raise HTTPException(400, "The uploaded file is not a valid Excel workbook")
    return safe_name, extension


def _store_source(
    *,
    project_id: str,
    entity_type: str,
    entity_id: str,
    filename: str,
    raw_bytes: bytes,
) -> str:
    target_dir = os.path.join(site_budget_dir(project_id), entity_type)
    os.makedirs(target_dir, exist_ok=True)
    extension = Path(filename).suffix.lower()
    target = os.path.join(target_dir, f"{entity_id}{extension}")
    with open(target, "wb") as handle:
        handle.write(raw_bytes)
    return target


def _header_key(value: Any) -> str:
    return " ".join(re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).split())


def _find_header_column(headers: list[str], candidates: Iterable[str]) -> int:
    normalized_candidates = [_header_key(item) for item in candidates]
    for index, header in enumerate(headers):
        if not header:
            continue
        if any(candidate == header or candidate in header for candidate in normalized_candidates):
            return index
    return -1


def _excel_rows(raw_bytes: bytes, *, document_type: str) -> tuple[list[dict[str, Any]], dict[str, Any], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise HTTPException(503, "Excel extraction requires openpyxl") from error
    try:
        workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    except Exception as error:
        raise HTTPException(422, "The uploaded Excel workbook could not be read") from error

    extracted: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    header: dict[str, Any] = {"currency": "", "document_number": ""}
    seen: set[tuple[str, str, str]] = set()
    for worksheet in workbook.worksheets:
        active_columns: dict[str, int] = {}
        blank_streak = 0
        for row_number, raw_row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = list(raw_row)
            keys = [_header_key(value) for value in values]
            description_index = _find_header_column(keys, ["description", "item description", "work description"])
            unit_index = _find_header_column(keys, ["unit", "uom"])
            amount_index = _find_header_column(keys, ["amount", "line amount", "current amount", "present amount"])
            if description_index >= 0 and (unit_index >= 0 or amount_index >= 0):
                active_columns = {
                    "item": _find_header_column(keys, ["s no", "item no", "boq item", "item"]),
                    "description": description_index,
                    "unit": unit_index,
                    "quantity": _find_header_column(keys, ["quantity", "qty", "contract quantity"]),
                    "rate": _find_header_column(keys, ["unit rate", "rate"]),
                    "amount": amount_index,
                    "previous_quantity": _find_header_column(keys, ["previous quantity", "previous qty"]),
                    "previous_amount": _find_header_column(keys, ["previous amount", "previous value"]),
                    "current_quantity": _find_header_column(keys, ["current quantity", "current qty", "this period quantity"]),
                    "current_amount": _find_header_column(keys, ["current amount", "present amount", "this period amount"]),
                    "cumulative_quantity": _find_header_column(keys, ["cumulative quantity", "total to date quantity"]),
                    "cumulative_amount": _find_header_column(keys, ["cumulative amount", "total to date amount", "to date amount"]),
                }
                blank_streak = 0
                continue
            if not active_columns:
                joined = " ".join(str(value or "") for value in values)
                currency_match = re.search(r"\b(SAR|USD|AED|QAR|EUR|GBP)\b", joined, re.IGNORECASE)
                if currency_match and not header["currency"]:
                    header["currency"] = currency_match.group(1).upper()
                continue

            def value_for(key: str) -> Any:
                index = active_columns.get(key, -1)
                return values[index] if index is not None and 0 <= index < len(values) else None

            description = str(value_for("description") or "").strip()
            item_number = str(value_for("item") or "").strip()
            unit = normalize_unit(value_for("unit"))
            if not description:
                blank_streak += 1
                if blank_streak > 8:
                    active_columns = {}
                continue
            blank_streak = 0
            description_key = normalize_description(description)
            if len(description_key) < 3 or description_key in {"DESCRIPTION", "TOTAL", "PAGE TOTAL"}:
                continue
            key = (worksheet.title, item_number, description_key)
            if key in seen:
                continue
            seen.add(key)
            quantity = number(value_for("quantity"))
            rate = number(value_for("rate"))
            amount = number(value_for("amount"))
            if document_type == "boq":
                if quantity <= 0 and rate <= 0 and amount <= 0:
                    continue
                extracted.append(
                    {
                        "line_id": f"line_{uuid4().hex}",
                        "item_number": item_number,
                        "description": description,
                        "category": worksheet.title,
                        "unit": unit,
                        "contract_qty": quantity,
                        "contract_unit_rate": rate,
                        "contract_amount": amount or quantity * rate,
                        "source_sheet": worksheet.title,
                        "source_row": row_number,
                        "confidence": 0.86,
                        "warnings": [],
                    }
                )
            else:
                current_quantity = number(value_for("current_quantity")) or quantity
                current_amount = number(value_for("current_amount")) or amount
                cumulative_quantity = number(value_for("cumulative_quantity"))
                cumulative_amount = number(value_for("cumulative_amount"))
                if current_quantity <= 0 and current_amount <= 0 and cumulative_amount <= 0:
                    continue
                extracted.append(
                    {
                        "line_id": f"line_{uuid4().hex}",
                        "item_number": item_number,
                        "description": description,
                        "unit": unit,
                        "contract_qty": quantity,
                        "claimed_unit_rate": rate,
                        "previous_certified_qty_source": number(value_for("previous_quantity")),
                        "previous_certified_amount_source": number(value_for("previous_amount")),
                        "current_claimed_qty": current_quantity,
                        "current_claimed_amount": current_amount or current_quantity * rate,
                        "source_cumulative_qty": cumulative_quantity,
                        "source_cumulative_amount": cumulative_amount,
                        "source_sheet": worksheet.title,
                        "source_row": row_number,
                        "confidence": 0.82,
                        "warnings": [],
                    }
                )
    if not extracted:
        warnings.append(
            {
                "code": "manual_line_review_required",
                "message": "No reliable line-item table was detected. Add or correct rows during review.",
            }
        )
    return extracted, header, warnings


def extract_budget_document(
    raw_bytes: bytes, *, filename: str, document_type: str
) -> tuple[list[dict[str, Any]], dict[str, Any], list[dict[str, Any]], str]:
    _, extension = _validate_upload(filename, raw_bytes)
    if extension in {".xlsx", ".xlsm"}:
        lines, header, warnings = _excel_rows(raw_bytes, document_type=document_type)
        return lines, header, warnings, "excel"

    material_type = "boq" if document_type == "boq" else "progress_invoice"
    pages, extraction_method, page_warnings = extract_pdf_pages(
        raw_bytes,
        document_type_hint=material_type,
        filename_hint=filename,
    )
    raw_lines, line_warnings = extract_structured_lines(
        pages, document_type=material_type
    )
    raw_header = extract_material_header("\n\f\n".join(pages), material_type)
    lines: list[dict[str, Any]] = []
    if document_type == "boq":
        for line in raw_lines:
            quantity = number(line.get("planned_qty"))
            rate = number(line.get("contract_unit_rate"))
            amount = number(line.get("line_amount")) or quantity * rate
            lines.append(
                {
                    **line,
                    "contract_qty": quantity,
                    "contract_unit_rate": rate,
                    "contract_amount": amount,
                    "category": str(line.get("category") or ""),
                }
            )
    else:
        for line in raw_lines:
            cumulative_qty = number(line.get("certified_qty"))
            cumulative_amount = number(line.get("certified_value"))
            lines.append(
                {
                    **line,
                    "contract_qty": number(line.get("contract_qty")),
                    "claimed_unit_rate": number(line.get("certified_unit_rate")),
                    "current_claimed_qty": 0.0,
                    "current_claimed_amount": 0.0,
                    "source_cumulative_qty": cumulative_qty,
                    "source_cumulative_amount": cumulative_amount,
                    "source_cumulative_percent": number(line.get("certified_percent")),
                }
            )
    warnings = [*page_warnings, *line_warnings]
    return lines, raw_header, warnings, extraction_method


def normalize_boq_lines(lines: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for index, raw in enumerate(lines):
        line = dict(raw or {})
        description = str(line.get("description") or "").strip()
        if not description:
            continue
        quantity = max(0.0, number(line.get("contract_qty", line.get("planned_qty"))))
        rate = max(0.0, number(line.get("contract_unit_rate")))
        amount = number(line.get("contract_amount", line.get("line_amount")))
        calculated_amount = quantity * rate
        amount_was_recalculated = calculated_amount > 0 and (
            amount <= 0
            or abs(amount - calculated_amount) > max(1.0, calculated_amount * 0.05)
        )
        if amount_was_recalculated:
            amount = calculated_amount
        line_warnings = [str(value) for value in (line.get("warnings") or [])]
        if amount_was_recalculated:
            line_warnings.append(
                "Amount was recalculated from quantity x unit rate during review."
            )
        output.append(
            {
                "line_id": str(line.get("line_id") or f"line_{uuid4().hex}"),
                "boq_item_id": str(line.get("boq_item_id") or "").strip(),
                "item_number": str(line.get("item_number") or index + 1).strip(),
                "description": description,
                "normalized_description": normalize_description(description),
                "category": str(line.get("category") or "Uncategorized").strip(),
                "unit": normalize_unit(line.get("unit")),
                "contract_qty": rounded(quantity),
                "contract_unit_rate": rounded(rate),
                "contract_amount": rounded(amount),
                "activity_id": str(line.get("activity_id") or "").strip(),
                "activity_ids": [
                    str(value).strip()
                    for value in (line.get("activity_ids") or [])
                    if str(value).strip()
                ],
                "zone": str(line.get("zone") or "").strip(),
                "work_category": str(line.get("work_category") or "").strip(),
                "material_id": str(line.get("material_id") or "").strip(),
                "source_page": int(number(line.get("source_page"))) or None,
                "source_sheet": str(line.get("source_sheet") or ""),
                "source_row": int(number(line.get("source_row"))) or None,
                "confidence": min(1.0, max(0.0, number(line.get("confidence")))),
                "warnings": line_warnings,
            }
        )
    return output


def normalize_invoice_lines(lines: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for index, raw in enumerate(lines):
        line = dict(raw or {})
        description = str(line.get("description") or "").strip()
        if not description:
            continue
        current_qty = max(0.0, number(line.get("current_claimed_qty")))
        rate = max(0.0, number(line.get("claimed_unit_rate")))
        current_amount = max(0.0, number(line.get("current_claimed_amount")))
        if current_amount <= 0 and current_qty > 0 and rate > 0:
            current_amount = current_qty * rate
        output.append(
            {
                "line_id": str(line.get("line_id") or f"line_{uuid4().hex}"),
                "boq_item_id": str(line.get("boq_item_id") or "").strip(),
                "item_number": str(line.get("item_number") or index + 1).strip(),
                "description": description,
                "normalized_description": normalize_description(description),
                "unit": normalize_unit(line.get("unit")),
                "claimed_unit_rate": rounded(rate),
                "current_claimed_qty": rounded(current_qty),
                "current_claimed_amount": rounded(current_amount),
                "source_cumulative_qty": rounded(line.get("source_cumulative_qty")),
                "source_cumulative_amount": rounded(line.get("source_cumulative_amount")),
                "source_cumulative_percent": rounded(line.get("source_cumulative_percent")),
                "manual_verified_percent": (
                    min(100.0, max(0.0, number(line.get("manual_verified_percent"))))
                    if line.get("manual_verified_percent") not in {None, ""}
                    else None
                ),
                "source_page": int(number(line.get("source_page"))) or None,
                "source_sheet": str(line.get("source_sheet") or ""),
                "source_row": int(number(line.get("source_row"))) or None,
                "confidence": min(1.0, max(0.0, number(line.get("confidence")))),
                "warnings": [str(value) for value in (line.get("warnings") or [])],
            }
        )
    return output


def _match_score(description: str, candidate: str) -> float:
    left = normalize_description(description)
    right = normalize_description(candidate)
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    overlap = len(left_tokens & right_tokens) / max(len(left_tokens | right_tokens), 1)
    sequence = SequenceMatcher(None, left, right).ratio()
    return max(sequence, overlap)


def _active_boq(project_id: str) -> Optional[dict[str, Any]]:
    return budget_boqs_collection.find_one(
        {"project_id": project_id, "is_active": True}, sort=[("version", -1)]
    )


def _active_boq_items(project_id: str) -> list[dict[str, Any]]:
    boq = _active_boq(project_id)
    if not boq:
        return []
    return list(
        budget_boq_items_collection.find(
            {"project_id": project_id, "boq_id": boq["boq_id"]}, {"_id": 0}
        ).sort([("category", 1), ("item_number", 1)])
    )


def _boq_summary(lines: Iterable[dict[str, Any]]) -> dict[str, Any]:
    values = list(lines)
    categories = {str(line.get("category") or "Uncategorized") for line in values}
    original_amount = sum(number(line.get("contract_amount")) for line in values)
    unmapped = sum(
        1
        for line in values
        if not str(line.get("activity_id") or "").strip()
        and not (line.get("activity_ids") or [])
    )
    return {
        "line_count": len(values),
        "category_count": len(categories),
        "original_contract_amount": rounded(original_amount),
        "unmapped_line_count": unmapped,
    }


def _next_boq_version(project_id: str) -> int:
    latest = budget_boqs_collection.find_one(
        {"project_id": project_id}, sort=[("version", -1)]
    )
    return int((latest or {}).get("version") or 0) + 1


def should_reuse_uploaded_boq(existing: Optional[dict[str, Any]]) -> bool:
    """Reuse a draft duplicate, but let an active source start a new revision."""
    if not existing:
        return False
    return not existing.get("is_active") and str(existing.get("status") or "") in {
        "needs_review",
        "reviewed",
    }


def upload_boq(
    *,
    project_ref: str,
    filename: str,
    raw_bytes: bytes,
    revision: str,
    currency: str,
    user: AuthenticatedUser,
) -> dict[str, Any]:
    safe_name, _ = _validate_upload(filename, raw_bytes)
    project = resolve_project(project_ref)
    project_id = project["project_id"]
    digest = hashlib.sha256(raw_bytes).hexdigest()
    existing = budget_boqs_collection.find_one(
        {"project_id": project_id, "source_sha256": digest},
        sort=[("version", -1)],
    )
    if should_reuse_uploaded_boq(existing):
        logger.info(
            "Budget BOQ duplicate upload reused",
            extra={
                "project_id": project_id,
                "boq_id": existing.get("boq_id"),
                "status": existing.get("status"),
            },
        )
        return {
            "status": "duplicate",
            "boq": get_boq(project_id, existing["boq_id"]),
        }
    extracted_lines, extracted_header, warnings, method = extract_budget_document(
        raw_bytes, filename=safe_name, document_type="boq"
    )
    resolved_revision = str(
        revision
        or extracted_header.get("revision")
        or revision_from_filename(safe_name)
    ).strip()
    resolved_currency = str(
        project_currency_code(project)
        or currency
        or extracted_header.get("currency")
        or ""
    ).strip().upper()
    extracted_header = {
        **dict(extracted_header or {}),
        "revision": resolved_revision,
        "currency": resolved_currency,
    }
    boq_id = f"boq_{uuid4().hex}"
    now = utc_now()
    document = {
        "boq_id": boq_id,
        "project_id": project_id,
        "site_name": project["site_name"],
        "floorplan_id": project["floorplan_id"],
        "version": _next_boq_version(project_id),
        "revision": resolved_revision,
        "currency": resolved_currency,
        "status": "needs_review",
        "is_active": False,
        "original_filename": safe_name,
        "source_sha256": digest,
        "extraction_method": method,
        "extraction_warnings": warnings,
        "extracted_header": extracted_header,
        "extracted_lines": normalize_boq_lines(extracted_lines),
        "reviewed_header": {},
        "reviewed_lines": [],
        "summary": _boq_summary(normalize_boq_lines(extracted_lines)),
        "uploaded_at": now,
        "uploaded_by_user_id": user.user_id,
        "uploaded_by_email": user.email,
        "created_at": now,
        "updated_at": now,
    }
    document["storage_path"] = _store_source(
        project_id=project_id,
        entity_type="boq",
        entity_id=boq_id,
        filename=safe_name,
        raw_bytes=raw_bytes,
    )
    budget_boqs_collection.insert_one(document)
    logger.info(
        "Budget BOQ uploaded for review",
        extra={
            "project_id": project_id,
            "boq_id": boq_id,
            "version": document["version"],
            "line_count": len(document["extracted_lines"]),
            "repeated_source": bool(existing),
        },
    )
    _audit(
        project_id=project_id,
        entity_type="boq",
        entity_id=boq_id,
        event_type="uploaded",
        user=user,
        details={"filename": safe_name, "line_count": len(document["extracted_lines"])},
    )
    return {"status": "needs_review", "boq": _public(document)}


def material_boq_to_budget_payload(
    document: dict[str, Any],
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    """Convert a confirmed Materials BOQ into a reviewable Budget BOQ payload."""
    source_header = dict(document.get("confirmed_header") or {})
    source_lines = [dict(line) for line in document.get("confirmed_lines") or []]
    refreshed_count = 0
    extracted_text = str(document.get("extracted_text") or "")
    if int(document.get("processing_version") or 0) < 10 and extracted_text:
        fresh_lines, _ = extract_structured_lines(
            extracted_text.split("\n\f\n"), document_type="boq"
        )
        fresh_by_key: dict[tuple[int, str], list[dict[str, Any]]] = defaultdict(list)
        for candidate in fresh_lines:
            key = (
                int(number(candidate.get("source_page"))),
                str(candidate.get("item_number") or "").strip(),
            )
            fresh_by_key[key].append(candidate)
        for line in source_lines:
            key = (
                int(number(line.get("source_page"))),
                str(line.get("item_number") or "").strip(),
            )
            candidates = fresh_by_key.get(key) or []
            if not candidates:
                continue
            candidate = max(
                candidates,
                key=lambda item: _match_score(
                    str(line.get("description") or ""),
                    str(item.get("description") or ""),
                ),
            )
            if (
                _match_score(
                    str(line.get("description") or ""),
                    str(candidate.get("description") or ""),
                )
                < 0.75
            ):
                continue
            previous = (
                rounded(line.get("planned_qty")),
                rounded(line.get("contract_unit_rate")),
                rounded(line.get("line_amount")),
            )
            replacement = (
                rounded(candidate.get("planned_qty")),
                rounded(candidate.get("contract_unit_rate")),
                rounded(candidate.get("line_amount")),
            )
            if min(replacement) <= 0 or replacement == previous:
                continue
            line.update(
                planned_qty=replacement[0],
                contract_unit_rate=replacement[1],
                line_amount=replacement[2],
                unit=candidate.get("unit") or line.get("unit"),
                warnings=list(candidate.get("warnings") or line.get("warnings") or []),
                confidence=candidate.get("confidence", line.get("confidence")),
            )
            refreshed_count += 1
    converted = normalize_boq_lines(
        [
            {
                **dict(line),
                "line_id": str(line.get("line_id") or f"line_{uuid4().hex}"),
                "contract_qty": number(line.get("planned_qty")),
                "contract_unit_rate": number(line.get("contract_unit_rate")),
                "contract_amount": number(line.get("line_amount"))
                or number(line.get("planned_qty"))
                * number(line.get("contract_unit_rate")),
                "material_id": str(line.get("material_id") or "").strip(),
            }
            for line in source_lines
        ]
    )
    zero_value_count = sum(
        1 for line in converted if number(line.get("contract_amount")) <= 0
    )
    warnings: list[dict[str, Any]] = [
        {
            "code": "imported_from_materials",
            "message": (
                "Imported from the confirmed Materials BOQ. Review quantities, rates, "
                "amounts, and activity mappings before activation."
            ),
        }
    ]
    if zero_value_count:
        warnings.append(
            {
                "code": "missing_priced_lines",
                "message": (
                    f"{zero_value_count} imported line(s) have no priced amount and require review."
                ),
            }
        )
    if refreshed_count:
        warnings.append(
            {
                "code": "financial_columns_reparsed",
                "message": (
                    f"{refreshed_count} line(s) used the corrected PDF quantity, rate, "
                    "and amount column layout. Confirm them before activation."
                ),
            }
        )
    return source_header, converted, warnings


def import_material_boq(
    *,
    project_ref: str,
    material_document_id: str,
    user: AuthenticatedUser,
) -> dict[str, Any]:
    project = resolve_project(project_ref)
    project_id = project["project_id"]
    source = material_documents_collection.find_one(
        {
            "project_id": project_id,
            "document_id": str(material_document_id or "").strip(),
            "document_type": "boq",
        }
    )
    if not source:
        raise HTTPException(404, "Materials BOQ not found")
    if source.get("status") != "confirmed":
        raise HTTPException(
            409, "Only the current confirmed Materials BOQ can be used in Budget"
        )

    extracted_header, extracted_lines, warnings = material_boq_to_budget_payload(
        source
    )
    if not extracted_lines:
        raise HTTPException(
            422, "The confirmed Materials BOQ has no line items available to import"
        )
    revision = str(
        source.get("revision")
        or extracted_header.get("revision")
        or revision_from_filename(source.get("original_filename"))
    ).strip()
    currency = str(
        project_currency_code(project)
        or source.get("currency")
        or extracted_header.get("currency")
        or ""
    ).strip().upper()
    extracted_header = {
        **extracted_header,
        "revision": revision,
        "currency": currency,
    }

    existing = budget_boqs_collection.find_one(
        {
            "project_id": project_id,
            "source_material_document_id": source["document_id"],
        }
    )
    if not existing and source.get("source_sha256"):
        existing = budget_boqs_collection.find_one(
            {"project_id": project_id, "source_sha256": source["source_sha256"]}
        )
    if existing:
        update = {
            "source_type": "materials_boq",
            "source_material_document_id": source["document_id"],
            "source_material_confirmed_at": source.get("confirmed_at"),
            "updated_at": utc_now(),
        }
        if existing.get("status") == "needs_review" and not existing.get("is_active"):
            update.update(
                revision=revision,
                currency=currency,
                extraction_method="materials_confirmed_boq",
                extraction_warnings=warnings,
                extracted_header=extracted_header,
                extracted_lines=extracted_lines,
                summary=_boq_summary(extracted_lines),
            )
        budget_boqs_collection.update_one(
            {"project_id": project_id, "boq_id": existing["boq_id"]},
            {"$set": update},
        )
        return {
            "status": "already_imported",
            "boq": get_boq(project_id, existing["boq_id"]),
        }

    boq_id = f"boq_{uuid4().hex}"
    now = utc_now()
    document = {
        "boq_id": boq_id,
        "project_id": project_id,
        "site_name": project["site_name"],
        "floorplan_id": project["floorplan_id"],
        "version": _next_boq_version(project_id),
        "revision": revision,
        "currency": currency,
        "status": "needs_review",
        "is_active": False,
        "original_filename": str(source.get("original_filename") or "Materials BOQ"),
        "source_sha256": str(source.get("source_sha256") or ""),
        "source_type": "materials_boq",
        "source_material_document_id": source["document_id"],
        "source_material_confirmed_at": source.get("confirmed_at"),
        "extraction_method": "materials_confirmed_boq",
        "extraction_warnings": warnings,
        "extracted_header": extracted_header,
        "extracted_lines": extracted_lines,
        "reviewed_header": {},
        "reviewed_lines": [],
        "summary": _boq_summary(extracted_lines),
        "uploaded_at": now,
        "uploaded_by_user_id": user.user_id,
        "uploaded_by_email": user.email,
        "created_at": now,
        "updated_at": now,
    }
    budget_boqs_collection.insert_one(document)
    _audit(
        project_id=project_id,
        entity_type="boq",
        entity_id=boq_id,
        event_type="imported_from_materials",
        user=user,
        details={
            "material_document_id": source["document_id"],
            "filename": document["original_filename"],
            "line_count": len(extracted_lines),
        },
    )
    return {"status": "needs_review", "boq": _public(document)}


def _materials_boq_sources(project: dict[str, Any]) -> list[dict[str, Any]]:
    project_id = project["project_id"]
    project_currency = project_currency_code(project)
    sources = material_documents_collection.find(
        {"project_id": project_id, "document_type": "boq", "status": "confirmed"}
    ).sort([("confirmed_at", -1), ("uploaded_at", -1)])
    values: list[dict[str, Any]] = []
    for source in sources:
        existing = budget_boqs_collection.find_one(
            {
                "project_id": project_id,
                "source_material_document_id": source.get("document_id"),
            }
        )
        if not existing and source.get("source_sha256"):
            existing = budget_boqs_collection.find_one(
                {
                    "project_id": project_id,
                    "source_sha256": source.get("source_sha256"),
                }
            )
        lines = list(source.get("confirmed_lines") or [])
        contract_amount = sum(
            number(line.get("line_amount"))
            or number(line.get("planned_qty"))
            * number(line.get("contract_unit_rate"))
            for line in lines
        )
        header = dict(source.get("confirmed_header") or {})
        values.append(
            {
                "document_id": source.get("document_id"),
                "original_filename": source.get("original_filename") or "Materials BOQ",
                "revision": source.get("revision")
                or header.get("revision")
                or revision_from_filename(source.get("original_filename")),
                "currency": project_currency
                or source.get("currency")
                or header.get("currency")
                or "",
                "line_count": len(lines),
                "contract_amount": rounded(contract_amount),
                "confirmed_at": source.get("confirmed_at"),
                "source_sha256": source.get("source_sha256") or "",
                "is_imported": bool(existing),
                "budget_boq_id": (existing or {}).get("boq_id") or "",
                "budget_status": (existing or {}).get("status") or "",
                "is_active_in_budget": bool((existing or {}).get("is_active")),
            }
        )
    return public_value(values)


def list_boqs(project_ref: str) -> list[dict[str, Any]]:
    project = resolve_project(project_ref)
    return [
        _public(item)
        for item in budget_boqs_collection.find(
            {"project_id": project["project_id"]}
        ).sort([("version", -1)])
    ]


def get_boq(project_ref: str, boq_id: str) -> dict[str, Any]:
    project = resolve_project(project_ref)
    document = budget_boqs_collection.find_one(
        {"project_id": project["project_id"], "boq_id": boq_id}
    )
    if not document:
        raise HTTPException(404, "BOQ revision not found")
    result = _public(document)
    configured_currency = project_currency_code(project)
    if configured_currency:
        result["currency"] = configured_currency
    if not str(result.get("revision") or "").strip():
        result["revision"] = revision_from_filename(result.get("original_filename"))
    result["extracted_header"] = {
        **dict(result.get("extracted_header") or {}),
        "revision": result.get("revision") or "",
        "currency": result.get("currency") or "",
    }
    return result


def review_boq(
    *,
    project_ref: str,
    boq_id: str,
    header: dict[str, Any],
    lines: list[dict[str, Any]],
    note: str,
    user: AuthenticatedUser,
) -> dict[str, Any]:
    project = resolve_project(project_ref)
    project_id = project["project_id"]
    document = budget_boqs_collection.find_one(
        {"project_id": project_id, "boq_id": boq_id}
    )
    if not document:
        raise HTTPException(404, "BOQ revision not found")
    if document.get("is_active") is True:
        logger.warning(
            "Active Budget BOQ edit rejected",
            extra={"project_id": project_id, "boq_id": boq_id},
        )
        raise HTTPException(409, "Upload a new BOQ revision instead of editing the active revision")
    normalized = normalize_boq_lines(lines)
    if not normalized:
        raise HTTPException(422, "At least one BOQ line is required")
    now = utc_now()
    reviewed_header = {
        **dict(document.get("extracted_header") or {}),
        **dict(header or {}),
    }
    resolved_revision = str(
        reviewed_header.get("revision")
        or document.get("revision")
        or revision_from_filename(document.get("original_filename"))
    ).strip()
    resolved_currency = str(
        project_currency_code(project)
        or reviewed_header.get("currency")
        or document.get("currency")
        or ""
    ).strip().upper()
    reviewed_header["revision"] = resolved_revision
    reviewed_header["currency"] = resolved_currency
    update = {
        "reviewed_header": reviewed_header,
        "reviewed_lines": normalized,
        "review_note": str(note or "").strip(),
        "revision": resolved_revision,
        "currency": resolved_currency,
        "status": "reviewed",
        "summary": _boq_summary(normalized),
        "reviewed_at": now,
        "reviewed_by_user_id": user.user_id,
        "reviewed_by_email": user.email,
        "updated_at": now,
    }
    budget_boqs_collection.update_one(
        {"project_id": project_id, "boq_id": boq_id}, {"$set": update}
    )
    logger.info(
        "Budget BOQ review saved",
        extra={
            "project_id": project_id,
            "boq_id": boq_id,
            "line_count": len(normalized),
        },
    )
    _audit(
        project_id=project_id,
        entity_type="boq",
        entity_id=boq_id,
        event_type="reviewed",
        user=user,
        details={"line_count": len(normalized), "note": str(note or "").strip()},
    )
    return get_boq(project_id, boq_id)


def _reuse_boq_item_ids(
    project_id: str, lines: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    active_items = _active_boq_items(project_id)
    by_item: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    by_description: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for item in active_items:
        by_item[(str(item.get("item_number") or ""), str(item.get("unit") or ""))].append(item)
        by_description[(str(item.get("normalized_description") or ""), str(item.get("unit") or ""))].append(item)
    output: list[dict[str, Any]] = []
    for line in lines:
        item = dict(line)
        existing_id = str(item.get("boq_item_id") or "").strip()
        if not existing_id:
            candidates = by_item.get(
                (str(item.get("item_number") or ""), str(item.get("unit") or "")), []
            )
            if len(candidates) != 1:
                candidates = by_description.get(
                    (str(item.get("normalized_description") or ""), str(item.get("unit") or "")), []
                )
            existing_id = (
                str(candidates[0].get("boq_item_id") or "")
                if len(candidates) == 1
                else ""
            )
        item["boq_item_id"] = existing_id or f"boq_item_{uuid4().hex}"
        output.append(item)
    return output


def activate_boq(
    *, project_ref: str, boq_id: str, user: AuthenticatedUser
) -> dict[str, Any]:
    project = resolve_project(project_ref)
    project_id = project["project_id"]
    document = budget_boqs_collection.find_one(
        {"project_id": project_id, "boq_id": boq_id}
    )
    if not document:
        raise HTTPException(404, "BOQ revision not found")
    lines = normalize_boq_lines(
        document.get("reviewed_lines") or document.get("extracted_lines") or []
    )
    if not lines or document.get("status") not in {"reviewed", "active"}:
        raise HTTPException(409, "Review and confirm the BOQ lines before activation")
    lines = _reuse_boq_item_ids(project_id, lines)
    now = utc_now()
    budget_boqs_collection.update_many(
        {"project_id": project_id, "is_active": True},
        {"$set": {"is_active": False, "status": "superseded", "superseded_at": now}},
    )
    budget_boq_items_collection.delete_many({"project_id": project_id, "boq_id": boq_id})
    budget_boq_items_collection.insert_many(
        [
            {
                **line,
                "project_id": project_id,
                "site_name": project["site_name"],
                "floorplan_id": project["floorplan_id"],
                "boq_id": boq_id,
                "boq_version": document.get("version"),
                "currency": document.get("currency") or "",
                "activated_at": now,
            }
            for line in lines
        ]
    )
    summary = _boq_summary(lines)
    budget_boqs_collection.update_one(
        {"project_id": project_id, "boq_id": boq_id},
        {
            "$set": {
                "is_active": True,
                "status": "active",
                "reviewed_lines": lines,
                "summary": summary,
                "activated_at": now,
                "activated_by_user_id": user.user_id,
                "activated_by_email": user.email,
                "updated_at": now,
            }
        },
    )
    logger.info(
        "Budget BOQ activated",
        extra={
            "project_id": project_id,
            "boq_id": boq_id,
            "line_count": len(lines),
            "contract_amount": summary.get("original_contract_amount"),
        },
    )
    _audit(
        project_id=project_id,
        entity_type="boq",
        entity_id=boq_id,
        event_type="activated",
        user=user,
        details=summary,
    )
    rebuild_invoice_history(project_id, user=user, audit_event=False)
    return get_boq(project_id, boq_id)


def list_variations(project_ref: str) -> list[dict[str, Any]]:
    project = resolve_project(project_ref)
    return [
        _public(item)
        for item in budget_variations_collection.find(
            {"project_id": project["project_id"]}
        ).sort([("effective_date", -1), ("created_at", -1)])
    ]


def create_variation(
    *, project_ref: str, payload: dict[str, Any], user: AuthenticatedUser
) -> dict[str, Any]:
    project = resolve_project(project_ref)
    project_id = project["project_id"]
    boq = _active_boq(project_id)
    if not boq:
        raise HTTPException(409, "Activate a priced BOQ before adding a variation")
    boq_item_id = str(payload.get("boq_item_id") or "").strip()
    if boq_item_id and not budget_boq_items_collection.find_one(
        {"project_id": project_id, "boq_id": boq["boq_id"], "boq_item_id": boq_item_id}
    ):
        raise HTTPException(422, "Variation BOQ item was not found in the active revision")
    status = str(payload.get("status") or "approved").strip().lower()
    if status not in {"pending", "approved", "rejected", "voided"}:
        raise HTTPException(422, "Unsupported variation status")
    effective_date = parse_iso_date(
        payload.get("effective_date"), field="effective_date", required=status == "approved"
    )
    now = utc_now()
    document = {
        "variation_id": f"variation_{uuid4().hex}",
        "project_id": project_id,
        "site_name": project["site_name"],
        "boq_id": boq["boq_id"],
        "boq_item_id": boq_item_id,
        "reference": str(payload.get("reference") or "").strip(),
        "description": str(payload.get("description") or "").strip(),
        "quantity_delta": rounded(payload.get("quantity_delta")),
        "amount_delta": rounded(payload.get("amount_delta")),
        "rate_override": rounded(payload.get("rate_override")),
        "effective_date": effective_date,
        "approval_date": parse_iso_date(payload.get("approval_date"), field="approval_date"),
        "status": status,
        "note": str(payload.get("note") or "").strip(),
        "created_at": now,
        "created_by_user_id": user.user_id,
        "created_by_email": user.email,
        "updated_at": now,
    }
    budget_variations_collection.insert_one(document)
    _audit(
        project_id=project_id,
        entity_type="variation",
        entity_id=document["variation_id"],
        event_type="created",
        user=user,
        details={"status": status, "reference": document["reference"]},
    )
    if status == "approved":
        rebuild_invoice_history(project_id, user=user, audit_event=False)
    return _public(document)


def _next_invoice_sequence(project_id: str) -> int:
    latest = budget_invoices_collection.find_one(
        {"project_id": project_id}, sort=[("sequence", -1)]
    )
    return int((latest or {}).get("sequence") or 0) + 1


def upload_invoice(
    *,
    project_ref: str,
    filename: str,
    raw_bytes: bytes,
    invoice_number: str,
    invoice_date: str,
    billing_start_date: str,
    billing_end_date: str,
    billing_cutoff_date: str,
    currency: str,
    retention_percent: float,
    advance_recovery_percent: float,
    vat_percent: float,
    user: AuthenticatedUser,
) -> dict[str, Any]:
    safe_name, _ = _validate_upload(filename, raw_bytes)
    project = resolve_project(project_ref)
    project_id = project["project_id"]
    if not _active_boq(project_id):
        raise HTTPException(409, "Activate a priced BOQ before uploading invoices")
    digest = hashlib.sha256(raw_bytes).hexdigest()
    existing = budget_invoices_collection.find_one(
        {"project_id": project_id, "source_sha256": digest}
    )
    if existing:
        return {
            "status": "duplicate",
            "invoice": get_invoice(project_id, existing["invoice_id"]),
        }
    extracted_lines, extracted_header, warnings, method = extract_budget_document(
        raw_bytes, filename=safe_name, document_type="invoice"
    )
    resolved_number = str(
        invoice_number or extracted_header.get("document_number") or ""
    ).strip()
    resolved_invoice_date = parse_iso_date(
        invoice_date or extracted_header.get("document_date"), field="invoice_date"
    )
    resolved_start = parse_iso_date(billing_start_date, field="billing_start_date")
    resolved_end = parse_iso_date(billing_end_date, field="billing_end_date")
    resolved_cutoff = parse_iso_date(
        billing_cutoff_date or billing_end_date or invoice_date,
        field="billing_cutoff_date",
        required=True,
    )
    if resolved_start and resolved_end and resolved_start > resolved_end:
        raise HTTPException(422, "Billing start date must be on or before billing end date")
    invoice_id = f"invoice_{uuid4().hex}"
    now = utc_now()
    normalized_lines = normalize_invoice_lines(extracted_lines)
    header = {
        **dict(extracted_header or {}),
        "invoice_number": resolved_number,
        "invoice_date": resolved_invoice_date,
        "billing_start_date": resolved_start,
        "billing_end_date": resolved_end,
        "billing_cutoff_date": resolved_cutoff,
        "currency": str(
            project_currency_code(project)
            or currency
            or extracted_header.get("currency")
            or ""
        ).strip().upper(),
        "retention_percent": rounded(retention_percent),
        "advance_recovery_percent": rounded(advance_recovery_percent),
        "vat_percent": rounded(vat_percent),
    }
    document = {
        "invoice_id": invoice_id,
        "project_id": project_id,
        "site_name": project["site_name"],
        "floorplan_id": project["floorplan_id"],
        "sequence": _next_invoice_sequence(project_id),
        "invoice_number": resolved_number,
        "invoice_date": resolved_invoice_date,
        "billing_start_date": resolved_start,
        "billing_end_date": resolved_end,
        "billing_cutoff_date": resolved_cutoff,
        "currency": header["currency"],
        "status": "needs_review",
        "original_filename": safe_name,
        "source_sha256": digest,
        "extraction_method": method,
        "extraction_warnings": warnings,
        "extracted_header": header,
        "extracted_lines": normalized_lines,
        "reviewed_header": {},
        "reviewed_lines": [],
        "comparison": {},
        "verification_results": [],
        "verification_version": 0,
        "payments": [],
        "decision_history": [],
        "uploaded_at": now,
        "uploaded_by_user_id": user.user_id,
        "uploaded_by_email": user.email,
        "created_at": now,
        "updated_at": now,
    }
    document["storage_path"] = _store_source(
        project_id=project_id,
        entity_type="invoice",
        entity_id=invoice_id,
        filename=safe_name,
        raw_bytes=raw_bytes,
    )
    budget_invoices_collection.insert_one(document)
    _audit(
        project_id=project_id,
        entity_type="invoice",
        entity_id=invoice_id,
        event_type="uploaded",
        user=user,
        details={"filename": safe_name, "line_count": len(normalized_lines)},
    )
    return {"status": "needs_review", "invoice": _public(document)}


def list_invoices(project_ref: str) -> list[dict[str, Any]]:
    project = resolve_project(project_ref)
    return [
        _public(item)
        for item in budget_invoices_collection.find(
            {"project_id": project["project_id"]}
        ).sort([("billing_cutoff_date", -1), ("sequence", -1)])
    ]


def get_invoice(project_ref: str, invoice_id: str) -> dict[str, Any]:
    project = resolve_project(project_ref)
    document = budget_invoices_collection.find_one(
        {"project_id": project["project_id"], "invoice_id": invoice_id}
    )
    if not document:
        raise HTTPException(404, "Invoice/payment application not found")
    result = _public(document)
    configured_currency = project_currency_code(project)
    if configured_currency:
        result["currency"] = configured_currency
    result["extracted_header"] = {
        **dict(result.get("extracted_header") or {}),
        "currency": result.get("currency") or "",
    }
    return result


def review_invoice(
    *,
    project_ref: str,
    invoice_id: str,
    header: dict[str, Any],
    lines: list[dict[str, Any]],
    note: str,
    user: AuthenticatedUser,
) -> dict[str, Any]:
    project = resolve_project(project_ref)
    project_id = project["project_id"]
    document = budget_invoices_collection.find_one(
        {"project_id": project_id, "invoice_id": invoice_id}
    )
    if not document:
        raise HTTPException(404, "Invoice/payment application not found")
    if document.get("status") in {"certified", "paid", "rejected"}:
        raise HTTPException(409, "Certified, paid, or rejected invoices cannot be edited")
    normalized = normalize_invoice_lines(lines)
    if not normalized:
        raise HTTPException(422, "At least one invoice line is required")
    reviewed_header = {
        **dict(document.get("extracted_header") or {}),
        **dict(header or {}),
    }
    resolved_currency = str(
        project_currency_code(project)
        or reviewed_header.get("currency")
        or document.get("currency")
        or ""
    ).strip().upper()
    reviewed_header["currency"] = resolved_currency
    cutoff = parse_iso_date(
        reviewed_header.get("billing_cutoff_date") or document.get("billing_cutoff_date"),
        field="billing_cutoff_date",
        required=True,
    )
    start = parse_iso_date(
        reviewed_header.get("billing_start_date") or document.get("billing_start_date"),
        field="billing_start_date",
    )
    end = parse_iso_date(
        reviewed_header.get("billing_end_date") or document.get("billing_end_date"),
        field="billing_end_date",
    )
    if start and end and start > end:
        raise HTTPException(422, "Billing start date must be on or before billing end date")
    now = utc_now()
    update = {
        "reviewed_header": reviewed_header,
        "reviewed_lines": normalized,
        "review_note": str(note or "").strip(),
        "invoice_number": str(reviewed_header.get("invoice_number") or document.get("invoice_number") or "").strip(),
        "invoice_date": parse_iso_date(reviewed_header.get("invoice_date") or document.get("invoice_date"), field="invoice_date"),
        "billing_start_date": start,
        "billing_end_date": end,
        "billing_cutoff_date": cutoff,
        "currency": resolved_currency,
        "status": "reviewed",
        "reviewed_at": now,
        "reviewed_by_user_id": user.user_id,
        "reviewed_by_email": user.email,
        "updated_at": now,
    }
    budget_invoices_collection.update_one(
        {"project_id": project_id, "invoice_id": invoice_id}, {"$set": update}
    )
    _audit(
        project_id=project_id,
        entity_type="invoice",
        entity_id=invoice_id,
        event_type="reviewed",
        user=user,
        details={"line_count": len(normalized), "note": str(note or "").strip()},
    )
    rebuild_invoice_history(project_id, user=user, audit_event=False)
    return get_invoice(project_id, invoice_id)


def _effective_variations(
    project_id: str, cutoff: str
) -> dict[str, list[dict[str, Any]]]:
    query: dict[str, Any] = {"project_id": project_id, "status": "approved"}
    if cutoff:
        query["effective_date"] = {"$lte": cutoff}
    output: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for variation in budget_variations_collection.find(query).sort([("effective_date", 1)]):
        output[str(variation.get("boq_item_id") or "")].append(variation)
    return output


def _match_invoice_line(
    line: dict[str, Any], boq_items: list[dict[str, Any]]
) -> tuple[Optional[dict[str, Any]], float, str]:
    explicit = str(line.get("boq_item_id") or "").strip()
    if explicit:
        for item in boq_items:
            if str(item.get("boq_item_id") or "") == explicit:
                return item, 1.0, "boq_item_id"
    item_number = str(line.get("item_number") or "").strip()
    if item_number:
        candidates = [
            item for item in boq_items if str(item.get("item_number") or "").strip() == item_number
        ]
        if len(candidates) == 1:
            return candidates[0], 0.98, "item_number"
    scored = sorted(
        [(_match_score(line.get("description") or "", item.get("description") or ""), item) for item in boq_items],
        key=lambda entry: entry[0],
        reverse=True,
    )
    if scored and scored[0][0] >= 0.82 and (
        len(scored) == 1 or scored[0][0] - scored[1][0] >= 0.12
    ):
        return scored[0][1], rounded(scored[0][0]), "description"
    return None, rounded(scored[0][0] if scored else 0.0), "unmatched"


def _activity_evidence(
    project_id: str, cutoff: str, boq_items: list[dict[str, Any]]
) -> tuple[dict[str, dict[str, Any]], float]:
    try:
        comparison = build_baseline_comparison(
            project_id,
            as_of=date.fromisoformat(cutoff) if cutoff else None,
        )
    except Exception:
        comparison = None
    if not comparison:
        return {}, 0.0
    by_activity = {
        str(item.get("activity_id") or "").strip(): item
        for item in comparison.get("activities") or []
        if str(item.get("activity_id") or "").strip()
    }
    output: dict[str, dict[str, Any]] = {}
    for boq in boq_items:
        activity_ids = [
            str(value).strip()
            for value in [boq.get("activity_id"), *(boq.get("activity_ids") or [])]
            if str(value or "").strip()
        ]
        mapping_method = "activity_id" if activity_ids else ""
        mapping_confidence = 1.0 if activity_ids else 0.0
        if not activity_ids:
            scored = sorted(
                [
                    (
                        _match_score(
                            boq.get("description") or "",
                            activity.get("activity_name") or "",
                        ),
                        activity_id,
                    )
                    for activity_id, activity in by_activity.items()
                ],
                reverse=True,
            )
            if scored and scored[0][0] >= 0.88 and (
                len(scored) == 1 or scored[0][0] - scored[1][0] >= 0.12
            ):
                activity_ids = [scored[0][1]]
                mapping_method = "description"
                mapping_confidence = rounded(scored[0][0])
        activities = [by_activity[value] for value in activity_ids if value in by_activity]
        if not activities:
            continue
        actual_values = [number(item.get("actual_percent")) for item in activities]
        evidence = [
            evidence_item
            for activity in activities
            for evidence_item in (activity.get("evidence") or [])
            if str(evidence_item.get("status") or "") == "approved"
        ]
        output[str(boq.get("boq_item_id") or "")] = {
            "activity_ids": activity_ids,
            "mapping_method": mapping_method,
            "mapping_confidence": mapping_confidence,
            "physical_progress_percent": rounded(sum(actual_values) / len(actual_values)),
            "approved_evidence_count": len(evidence),
            "tour_evidence_count": sum(
                1
                for item in evidence
                if str(item.get("tour_id") or "").strip()
                and not str(item.get("tour_id") or "").startswith("manual:")
            ),
            "evidence": [
                {
                    "evidence_id": item.get("evidence_id"),
                    "tour_id": item.get("tour_id"),
                    "captured_at": item.get("captured_at"),
                    "approved_percent": item.get("approved_percent"),
                    "image_url": item.get("image_url"),
                    "source": item.get("review_source") or "tour",
                }
                for item in evidence[:20]
            ],
        }
    physical = number((comparison.get("summary") or {}).get("actual_percent"))
    return output, rounded(physical)


def _material_source_date(document: dict[str, Any]) -> str:
    candidates: list[Any] = []
    header = {
        **dict(document.get("extracted_header") or {}),
        **dict(document.get("confirmed_header") or {}),
    }
    for key in (
        "actual_delivery_date",
        "inspection_date",
        "document_date",
        "report_date",
        "delivery_date",
    ):
        candidates.extend([header.get(key), document.get(key)])
    for line in document.get("confirmed_lines") or []:
        for key in (
            "actual_delivery_date",
            "inspection_date",
            "document_date",
            "report_date",
            "delivery_date",
        ):
            candidates.append(line.get(key))
    valid_dates: list[str] = []
    for candidate in candidates:
        raw = str(candidate or "").strip()[:10]
        try:
            valid_dates.append(date.fromisoformat(raw).isoformat())
        except ValueError:
            continue
    return max(valid_dates) if valid_dates else ""


def material_documents_as_of(
    documents: Iterable[dict[str, Any]], cutoff: str
) -> tuple[list[dict[str, Any]], dict[str, list[str]]]:
    confirmed = [item for item in documents if item.get("status") == "confirmed"]
    if not cutoff:
        return confirmed, {}
    eligible: list[dict[str, Any]] = []
    undated_by_material: dict[str, list[str]] = defaultdict(list)
    for document in confirmed:
        document_type = str(document.get("document_type") or "")
        if document_type == "boq":
            eligible.append(document)
            continue
        source_date = _material_source_date(document)
        if source_date and source_date <= cutoff:
            eligible.append(document)
            continue
        if source_date:
            continue
        document_id = str(document.get("document_id") or "")
        for line in document.get("confirmed_lines") or []:
            material_id = str(line.get("linked_material_id") or "").strip()
            if material_id and document_id:
                undated_by_material[material_id].append(document_id)
    return eligible, {
        material_id: list(dict.fromkeys(document_ids))
        for material_id, document_ids in undated_by_material.items()
    }


def _material_evidence(
    project_id: str, boq_items: list[dict[str, Any]], cutoff: str
) -> dict[str, dict[str, Any]]:
    if cutoff:
        documents = list(
            material_documents_collection.find({"project_id": project_id}, {"_id": 0})
        )
        eligible, undated_by_material = material_documents_as_of(documents, cutoff)
        materials, _ = build_material_ledger(eligible)
    else:
        materials = list(
            project_materials_collection.find({"project_id": project_id}, {"_id": 0})
        )
        undated_by_material = {}
    output: dict[str, dict[str, Any]] = {}
    for boq in boq_items:
        explicit = str(boq.get("material_id") or "").strip()
        candidates = [item for item in materials if explicit and str(item.get("material_id") or "") == explicit]
        if not candidates:
            item_number = str(boq.get("item_number") or "").strip()
            candidates = [
                item for item in materials if item_number and str(item.get("boq_item_number") or "").strip() == item_number
            ]
        if not candidates:
            candidates = [
                item
                for item in materials
                if _match_score(boq.get("description") or "", item.get("description") or "") >= 0.94
                and normalize_unit(boq.get("unit")) == normalize_unit(item.get("unit"))
            ]
        if len(candidates) != 1:
            continue
        material = candidates[0]
        output[str(boq.get("boq_item_id") or "")] = {
            "material_id": material.get("material_id"),
            "unit": normalize_unit(material.get("unit")),
            "ordered_qty": rounded(material.get("ordered_qty")),
            "delivered_qty": rounded(material.get("delivered_qty")),
            "accepted_qty": rounded(material.get("accepted_qty")),
            "rejected_qty": rounded(material.get("rejected_qty")),
            "pending_inspection_qty": rounded(material.get("pending_inspection_qty")),
            "status": str(material.get("status") or ""),
            "source_document_ids": material.get("source_document_ids") or [],
            "evidence_cutoff_date": cutoff,
            "undated_document_ids": undated_by_material.get(
                str(material.get("material_id") or ""), []
            ),
        }
    return output


def calculate_invoice_line(
    *,
    invoice_line: dict[str, Any],
    boq_item: Optional[dict[str, Any]],
    variations: Iterable[dict[str, Any]],
    previous: dict[str, float],
    activity: Optional[dict[str, Any]],
    material: Optional[dict[str, Any]],
    match_confidence: float,
    match_method: str,
) -> dict[str, Any]:
    result = dict(invoice_line)
    reasons: list[str] = []
    if not boq_item:
        result.update(
            {
                "verification_status": "needs_review",
                "verification_reasons": ["Invoice line is not mapped to an active BOQ item."],
                "match_confidence": match_confidence,
                "match_method": match_method,
                "previous_certified_qty": 0.0,
                "previous_certified_amount": 0.0,
                "cumulative_claimed_qty": rounded(invoice_line.get("current_claimed_qty")),
                "cumulative_claimed_amount": rounded(invoice_line.get("current_claimed_amount")),
                "verified_cumulative_qty": 0.0,
                "verified_cumulative_amount": 0.0,
                "recommended_current_amount": 0.0,
                "variance_amount": rounded(invoice_line.get("current_claimed_amount")),
                "evidence": {"activity": {}, "material": {}},
            }
        )
        return result

    boq_item_id = str(boq_item.get("boq_item_id") or "")
    contract_qty = number(boq_item.get("contract_qty"))
    contract_rate = number(boq_item.get("contract_unit_rate"))
    contract_amount = number(boq_item.get("contract_amount")) or contract_qty * contract_rate
    revised_qty = contract_qty
    revised_amount = contract_amount
    revised_rate = contract_rate
    for variation in variations:
        revised_qty += number(variation.get("quantity_delta"))
        revised_amount += number(variation.get("amount_delta"))
        if number(variation.get("rate_override")) > 0:
            revised_rate = number(variation.get("rate_override"))
    if revised_amount <= 0 and revised_qty > 0 and revised_rate > 0:
        revised_amount = revised_qty * revised_rate

    previous_qty = max(0.0, number(previous.get("quantity")))
    previous_amount = max(0.0, number(previous.get("amount")))
    current_qty = max(0.0, number(invoice_line.get("current_claimed_qty")))
    current_amount = max(0.0, number(invoice_line.get("current_claimed_amount")))
    source_cumulative_qty = max(0.0, number(invoice_line.get("source_cumulative_qty")))
    source_cumulative_amount = max(0.0, number(invoice_line.get("source_cumulative_amount")))
    if current_qty <= 0 and source_cumulative_qty > 0:
        current_qty = max(source_cumulative_qty - previous_qty, 0.0)
    if current_amount <= 0 and source_cumulative_amount > 0:
        current_amount = max(source_cumulative_amount - previous_amount, 0.0)
    claimed_rate = number(invoice_line.get("claimed_unit_rate"))
    if current_amount <= 0 and current_qty > 0 and claimed_rate > 0:
        current_amount = current_qty * claimed_rate
    if current_qty <= 0 and current_amount > 0 and claimed_rate > 0:
        current_qty = current_amount / claimed_rate
    cumulative_qty = previous_qty + current_qty
    cumulative_amount = previous_amount + current_amount

    manual_percent = invoice_line.get("manual_verified_percent")
    activity_percent = number((activity or {}).get("physical_progress_percent"))
    supported_percent = (
        min(100.0, max(0.0, number(manual_percent)))
        if manual_percent is not None
        else min(100.0, max(0.0, activity_percent))
    )
    supported_qty = max(0.0, revised_qty * supported_percent / 100.0)
    supported_amount = max(0.0, revised_amount * supported_percent / 100.0)

    material_is_linked = material is not None
    material_unit_matches = material_is_linked and (
        not normalize_unit(boq_item.get("unit"))
        or normalize_unit(boq_item.get("unit")) == normalize_unit(material.get("unit"))
    )
    if material_is_linked and material_unit_matches and number(material.get("accepted_qty")) > 0:
        supported_qty = min(supported_qty, number(material.get("accepted_qty")))
        if revised_qty > 0:
            supported_amount = min(supported_amount, revised_amount * supported_qty / revised_qty)

    commercial_remaining = max(revised_amount - previous_amount, 0.0)
    evidence_remaining = max(supported_amount - previous_amount, 0.0)
    recommended = min(current_amount, commercial_remaining, evidence_remaining)

    status = "verified"
    if cumulative_qty > revised_qty + VERIFICATION_TOLERANCE or cumulative_amount > revised_amount + VERIFICATION_TOLERANCE:
        status = "overbilled"
        reasons.append("Cumulative claim exceeds the revised BOQ quantity or amount.")
    elif claimed_rate > 0 and revised_rate > 0 and abs(claimed_rate - revised_rate) > VERIFICATION_TOLERANCE:
        status = "mismatch"
        reasons.append("Claimed unit rate does not match the approved BOQ/variation rate.")
    elif supported_percent <= 0:
        status = "unverified"
        reasons.append("No approved activity/tour progress supports this line by the cutoff date.")
    elif previous_amount >= supported_amount - VERIFICATION_TOLERANCE and current_amount > 0:
        status = "duplicate"
        reasons.append("Previously certified value already uses the currently supported progress.")
    elif cumulative_amount > supported_amount + VERIFICATION_TOLERANCE:
        status = "mismatch"
        reasons.append("Cumulative claim is above verified physical progress.")
    if material_is_linked:
        if not material_unit_matches:
            status = "needs_review" if status == "verified" else status
            reasons.append("Material evidence uses an incompatible unit.")
        elif number(material.get("pending_inspection_qty")) > 0:
            status = "needs_review" if status == "verified" else status
            reasons.append("Related delivered material is still pending inspection.")
        elif number(material.get("accepted_qty")) + VERIFICATION_TOLERANCE < min(cumulative_qty, revised_qty):
            status = "needs_review" if status == "verified" else status
            reasons.append("Accepted material quantity is below the cumulative installed claim.")
        if material.get("undated_document_ids"):
            status = "needs_review" if status == "verified" else status
            reasons.append(
                "Related material evidence has no delivery/inspection source date for the billing cutoff."
            )
    if (
        activity
        and activity.get("mapping_method") == "description"
        and number(activity.get("mapping_confidence")) < 0.9
    ):
        status = "needs_review" if status == "verified" else status
        reasons.append("Description-based Activity mapping requires reviewer confirmation.")
    if match_method == "description" and match_confidence < 0.9:
        status = "needs_review" if status == "verified" else status
        reasons.append("Description-based BOQ mapping requires reviewer confirmation.")
    if not reasons and status == "verified":
        reasons.append("Claim is within revised BOQ limits and supported by approved progress evidence.")

    result.update(
        {
            "boq_item_id": boq_item_id,
            "boq_item_number": boq_item.get("item_number"),
            "boq_description": boq_item.get("description"),
            "match_confidence": rounded(match_confidence),
            "match_method": match_method,
            "contract_qty": rounded(contract_qty),
            "contract_unit_rate": rounded(contract_rate),
            "contract_amount": rounded(contract_amount),
            "revised_qty": rounded(revised_qty),
            "revised_unit_rate": rounded(revised_rate),
            "revised_amount": rounded(revised_amount),
            "previous_certified_qty": rounded(previous_qty),
            "previous_certified_amount": rounded(previous_amount),
            "current_claimed_qty": rounded(current_qty),
            "current_claimed_amount": rounded(current_amount),
            "cumulative_claimed_qty": rounded(cumulative_qty),
            "cumulative_claimed_amount": rounded(cumulative_amount),
            "verified_progress_percent": rounded(supported_percent),
            "verified_cumulative_qty": rounded(supported_qty),
            "verified_cumulative_amount": rounded(supported_amount),
            "recommended_current_amount": rounded(max(recommended, 0.0)),
            "variance_amount": rounded(max(cumulative_amount - supported_amount, 0.0)),
            "verification_status": status,
            "verification_reasons": reasons,
            "evidence": {"activity": activity or {}, "material": material or {}},
        }
    )
    return result


def _comparison_for_invoice(
    *,
    project_id: str,
    invoice: dict[str, Any],
    previous_totals: dict[str, dict[str, float]],
) -> tuple[dict[str, Any], list[dict[str, Any]], float]:
    boq = _active_boq(project_id)
    if not boq:
        raise HTTPException(409, "No active priced BOQ is available")
    boq_items = _active_boq_items(project_id)
    cutoff = str(invoice.get("billing_cutoff_date") or "")
    variations = _effective_variations(project_id, cutoff)
    activity_by_item, physical_percent = _activity_evidence(project_id, cutoff, boq_items)
    material_by_item = _material_evidence(project_id, boq_items, cutoff)
    invoice_lines = normalize_invoice_lines(
        invoice.get("reviewed_lines") or invoice.get("extracted_lines") or []
    )
    results: list[dict[str, Any]] = []
    for line in invoice_lines:
        matched, confidence, method = _match_invoice_line(line, boq_items)
        item_id = str((matched or {}).get("boq_item_id") or "")
        results.append(
            calculate_invoice_line(
                invoice_line=line,
                boq_item=matched,
                variations=variations.get(item_id, []),
                previous=previous_totals.get(item_id, {}),
                activity=activity_by_item.get(item_id),
                material=material_by_item.get(item_id),
                match_confidence=confidence,
                match_method=method,
            )
        )

    original_contract = sum(number(item.get("contract_amount")) for item in boq_items)
    approved_variation_amount = sum(
        number(item.get("amount_delta"))
        for variations_for_item in variations.values()
        for item in variations_for_item
    )
    revised_contract = original_contract + approved_variation_amount
    previous_amount = sum(number(item.get("previous_certified_amount")) for item in results)
    current_claimed = sum(number(item.get("current_claimed_amount")) for item in results)
    cumulative_claimed = previous_amount + current_claimed
    verified_cumulative = sum(number(item.get("verified_cumulative_amount")) for item in results)
    recommended_gross = sum(number(item.get("recommended_current_amount")) for item in results)
    header = {
        **dict(invoice.get("extracted_header") or {}),
        **dict(invoice.get("reviewed_header") or {}),
    }
    retention_percent = max(0.0, number(header.get("retention_percent")))
    advance_percent = max(0.0, number(header.get("advance_recovery_percent")))
    vat_percent = max(0.0, number(header.get("vat_percent")))
    retention_amount = current_claimed * retention_percent / 100.0
    advance_amount = current_claimed * advance_percent / 100.0
    net_before_vat = max(current_claimed - retention_amount - advance_amount, 0.0)
    vat_amount = net_before_vat * vat_percent / 100.0
    requested_payable = net_before_vat + vat_amount
    recommended_net = max(
        recommended_gross
        - recommended_gross * retention_percent / 100.0
        - recommended_gross * advance_percent / 100.0,
        0.0,
    )
    recommended_payable = recommended_net * (1 + vat_percent / 100.0)
    status_counts: dict[str, int] = defaultdict(int)
    for item in results:
        status_counts[str(item.get("verification_status") or "needs_review")] += 1
    comparison = {
        "original_contract_amount": rounded(original_contract),
        "approved_variation_amount": rounded(approved_variation_amount),
        "revised_contract_amount": rounded(revised_contract),
        "previous_certified_amount": rounded(previous_amount),
        "current_claimed_amount": rounded(current_claimed),
        "cumulative_claimed_amount": rounded(cumulative_claimed),
        "verified_cumulative_amount": rounded(verified_cumulative),
        "recommended_current_gross": rounded(recommended_gross),
        "retention_percent": rounded(retention_percent),
        "retention_amount": rounded(retention_amount),
        "advance_recovery_percent": rounded(advance_percent),
        "advance_recovery_amount": rounded(advance_amount),
        "vat_percent": rounded(vat_percent),
        "vat_amount": rounded(vat_amount),
        "requested_payable_amount": rounded(requested_payable),
        "recommended_payable_amount": rounded(recommended_payable),
        "remaining_contract_value": rounded(max(revised_contract - previous_amount, 0.0)),
        "current_claim_percentage": rounded(current_claimed / revised_contract * 100 if revised_contract else 0.0),
        "cumulative_billed_percentage": rounded(cumulative_claimed / revised_contract * 100 if revised_contract else 0.0),
        "verified_progress_percentage": rounded(verified_cumulative / revised_contract * 100 if revised_contract else 0.0),
        "project_physical_progress_percentage": rounded(physical_percent),
        "billing_progress_variance_percentage": rounded(
            cumulative_claimed / revised_contract * 100 - physical_percent
            if revised_contract
            else -physical_percent
        ),
        "variance_amount": rounded(max(cumulative_claimed - verified_cumulative, 0.0)),
        "status_counts": dict(status_counts),
        "flagged_line_count": sum(count for status, count in status_counts.items() if status != "verified"),
        "verified_line_count": status_counts.get("verified", 0),
        "calculated_at": utc_now(),
        "billing_cutoff_date": cutoff,
    }
    return comparison, results, physical_percent


def _invoice_sort_key(invoice: dict[str, Any]) -> tuple[str, str, int, str]:
    return (
        str(invoice.get("billing_cutoff_date") or "9999-12-31"),
        str(invoice.get("invoice_date") or "9999-12-31"),
        int(invoice.get("sequence") or 0),
        str(invoice.get("invoice_id") or ""),
    )


def rebuild_invoice_history(
    project_ref: str,
    *,
    user: Optional[AuthenticatedUser] = None,
    audit_event: bool = True,
) -> list[dict[str, Any]]:
    project = resolve_project(project_ref)
    project_id = project["project_id"]
    if not _active_boq(project_id):
        return []
    invoices = sorted(
        list(budget_invoices_collection.find({"project_id": project_id})),
        key=_invoice_sort_key,
    )
    previous: dict[str, dict[str, float]] = defaultdict(
        lambda: {"quantity": 0.0, "amount": 0.0}
    )
    updated: list[dict[str, Any]] = []
    now = utc_now()
    for invoice in invoices:
        comparison, results, _ = _comparison_for_invoice(
            project_id=project_id,
            invoice=invoice,
            previous_totals=previous,
        )
        current_status = str(invoice.get("status") or "needs_review")
        if current_status in INVOICE_ACTIVE_STATUSES or current_status == "rejected":
            next_status = current_status
        else:
            next_status = "needs_review"
        budget_invoices_collection.update_one(
            {"project_id": project_id, "invoice_id": invoice["invoice_id"]},
            {
                "$set": {
                    "comparison": comparison,
                    "verification_results": results,
                    "status": next_status,
                    "history_recalculated_at": now,
                    "updated_at": now,
                }
            },
        )
        if current_status in INVOICE_HISTORY_STATUSES:
            certified_lines = invoice.get("certified_lines") or []
            if not certified_lines:
                certified_lines = [
                    {
                        "boq_item_id": item.get("boq_item_id"),
                        "certified_current_qty": item.get("current_claimed_qty"),
                        "certified_current_amount": item.get("recommended_current_amount"),
                    }
                    for item in results
                ]
            for line in certified_lines:
                item_id = str(line.get("boq_item_id") or "")
                if not item_id:
                    continue
                previous[item_id]["quantity"] += number(line.get("certified_current_qty"))
                previous[item_id]["amount"] += number(line.get("certified_current_amount"))
        updated.append({**invoice, "comparison": comparison, "verification_results": results})
    if user and audit_event:
        _audit(
            project_id=project_id,
            entity_type="project_budget",
            entity_id=project_id,
            event_type="invoice_history_rebuilt",
            user=user,
            details={"invoice_count": len(updated)},
        )
    return [_public(item) for item in updated]


def verify_invoice(
    *, project_ref: str, invoice_id: str, user: AuthenticatedUser
) -> dict[str, Any]:
    project = resolve_project(project_ref)
    project_id = project["project_id"]
    document = budget_invoices_collection.find_one(
        {"project_id": project_id, "invoice_id": invoice_id}
    )
    if not document:
        raise HTTPException(404, "Invoice/payment application not found")
    if document.get("status") == "needs_review" and not document.get("reviewed_lines"):
        raise HTTPException(409, "Review the extracted invoice before verification")
    rebuild_invoice_history(project_id, user=user, audit_event=False)
    refreshed = budget_invoices_collection.find_one(
        {"project_id": project_id, "invoice_id": invoice_id}
    ) or document
    version = int(refreshed.get("verification_version") or 0) + 1
    run = {
        "verification_run_id": f"verification_{uuid4().hex}",
        "project_id": project_id,
        "invoice_id": invoice_id,
        "version": version,
        "comparison": refreshed.get("comparison") or {},
        "results": refreshed.get("verification_results") or [],
        "created_at": utc_now(),
        "created_by_user_id": user.user_id,
        "created_by_email": user.email,
    }
    budget_verification_runs_collection.insert_one(run)
    budget_invoices_collection.update_one(
        {"project_id": project_id, "invoice_id": invoice_id},
        {
            "$set": {
                "verification_version": version,
                "latest_verification_run_id": run["verification_run_id"],
                "verified_at": run["created_at"],
                "verified_by_user_id": user.user_id,
                "verified_by_email": user.email,
                "status": "verified",
                "updated_at": run["created_at"],
            }
        },
    )
    _audit(
        project_id=project_id,
        entity_type="invoice",
        entity_id=invoice_id,
        event_type="verified",
        user=user,
        details={
            "version": version,
            "flagged_line_count": (run["comparison"] or {}).get("flagged_line_count", 0),
        },
    )
    return get_invoice(project_id, invoice_id)


def decide_invoice(
    *,
    project_ref: str,
    invoice_id: str,
    action: str,
    note: str,
    certified_amount: Optional[float],
    user: AuthenticatedUser,
) -> dict[str, Any]:
    normalized_action = str(action or "").strip().lower()
    if normalized_action not in DECISION_ACTIONS:
        raise HTTPException(422, "Unsupported invoice decision")
    if normalized_action != "certify" and len(str(note or "").strip()) < 3:
        raise HTTPException(422, "A decision note is required")
    project = resolve_project(project_ref)
    project_id = project["project_id"]
    invoice = budget_invoices_collection.find_one(
        {"project_id": project_id, "invoice_id": invoice_id}
    )
    if not invoice:
        raise HTTPException(404, "Invoice/payment application not found")
    if normalized_action == "certify" and invoice.get("status") not in {
        "verified",
        "on_hold",
        "correction_requested",
    }:
        raise HTTPException(409, "Verify the invoice before certification")
    now = utc_now()
    status = {
        "certify": "certified",
        "hold": "on_hold",
        "request_correction": "correction_requested",
        "reject": "rejected",
    }[normalized_action]
    update: dict[str, Any] = {
        "status": status,
        "last_decision": normalized_action,
        "decision_note": str(note or "").strip(),
        "decided_at": now,
        "decided_by_user_id": user.user_id,
        "decided_by_email": user.email,
        "updated_at": now,
    }
    history_entry = {
        "action": normalized_action,
        "status": status,
        "note": str(note or "").strip(),
        "actor_user_id": user.user_id,
        "actor_email": user.email,
        "created_at": now,
    }
    if normalized_action == "certify":
        results = list(invoice.get("verification_results") or [])
        recommended = sum(number(item.get("recommended_current_amount")) for item in results)
        approved_gross = max(0.0, number(certified_amount) if certified_amount is not None else recommended)
        if approved_gross > recommended + VERIFICATION_TOLERANCE:
            raise HTTPException(422, "Certified amount cannot exceed the verified recommendation")
        ratio = approved_gross / recommended if recommended > 0 else 0.0
        certified_lines = [
            {
                "boq_item_id": item.get("boq_item_id"),
                "line_id": item.get("line_id"),
                "certified_current_qty": rounded(number(item.get("current_claimed_qty")) * ratio),
                "certified_current_amount": rounded(number(item.get("recommended_current_amount")) * ratio),
            }
            for item in results
            if item.get("boq_item_id")
        ]
        comparison = invoice.get("comparison") or {}
        retention_percent = number(comparison.get("retention_percent"))
        advance_percent = number(comparison.get("advance_recovery_percent"))
        vat_percent = number(comparison.get("vat_percent"))
        net = max(
            approved_gross
            - approved_gross * retention_percent / 100.0
            - approved_gross * advance_percent / 100.0,
            0.0,
        )
        payable = net * (1 + vat_percent / 100.0)
        update.update(
            {
                "certified_current_amount": rounded(approved_gross),
                "certified_payable_amount": rounded(payable),
                "certified_lines": certified_lines,
                "certified_at": now,
                "certified_by_user_id": user.user_id,
                "certified_by_email": user.email,
            }
        )
        history_entry["certified_current_amount"] = rounded(approved_gross)
        history_entry["certified_payable_amount"] = rounded(payable)
    budget_invoices_collection.update_one(
        {"project_id": project_id, "invoice_id": invoice_id},
        {"$set": update, "$push": {"decision_history": history_entry}},
    )
    _audit(
        project_id=project_id,
        entity_type="invoice",
        entity_id=invoice_id,
        event_type=normalized_action,
        user=user,
        details={"note": str(note or "").strip(), "certified_amount": certified_amount},
    )
    rebuild_invoice_history(project_id, user=user, audit_event=False)
    return get_invoice(project_id, invoice_id)


def calculate_payment_state(
    *, amount: Any, existing_paid: Any, certified_payable: Any
) -> tuple[float, str]:
    paid_amount = number(amount)
    if paid_amount <= 0:
        raise HTTPException(422, "Payment amount must be greater than zero")
    already_paid = number(existing_paid)
    certified_total = number(certified_payable)
    remaining_payable = max(certified_total - already_paid, 0.0)
    if remaining_payable <= VERIFICATION_TOLERANCE:
        raise HTTPException(409, "The certified invoice has already been paid in full")
    if paid_amount > remaining_payable + VERIFICATION_TOLERANCE:
        raise HTTPException(
            422,
            "Payment amount cannot exceed the remaining certified payable amount",
        )
    paid_total = rounded(already_paid + paid_amount)
    status = (
        "paid"
        if paid_total >= certified_total - VERIFICATION_TOLERANCE
        else "certified"
    )
    return paid_total, status


def record_payment(
    *,
    project_ref: str,
    invoice_id: str,
    amount: float,
    payment_date: str,
    reference: str,
    note: str,
    user: AuthenticatedUser,
) -> dict[str, Any]:
    project = resolve_project(project_ref)
    project_id = project["project_id"]
    invoice = budget_invoices_collection.find_one(
        {"project_id": project_id, "invoice_id": invoice_id}
    )
    if not invoice:
        raise HTTPException(404, "Invoice/payment application not found")
    if invoice.get("status") not in {"certified", "paid"}:
        raise HTTPException(409, "Only a certified invoice can receive a payment")
    resolved_date = parse_iso_date(payment_date, field="payment_date", required=True)
    existing_paid = sum(number(item.get("amount")) for item in invoice.get("payments") or [])
    certified_payable = number(invoice.get("certified_payable_amount"))
    paid_total, status = calculate_payment_state(
        amount=amount,
        existing_paid=existing_paid,
        certified_payable=certified_payable,
    )
    paid_amount = number(amount)
    payment = {
        "payment_id": f"payment_{uuid4().hex}",
        "amount": rounded(paid_amount),
        "payment_date": resolved_date,
        "reference": str(reference or "").strip(),
        "note": str(note or "").strip(),
        "recorded_at": utc_now(),
        "recorded_by_user_id": user.user_id,
        "recorded_by_email": user.email,
    }
    budget_invoices_collection.update_one(
        {"project_id": project_id, "invoice_id": invoice_id},
        {
            "$push": {"payments": payment},
            "$set": {
                "status": status,
                "paid_amount": paid_total,
                "last_payment_date": resolved_date,
                "payment_reference": payment["reference"],
                "updated_at": utc_now(),
            },
        },
    )
    _audit(
        project_id=project_id,
        entity_type="invoice",
        entity_id=invoice_id,
        event_type="payment_recorded",
        user=user,
        details={"payment_id": payment["payment_id"], "amount": paid_amount},
    )
    return get_invoice(project_id, invoice_id)


def get_budget_workspace(project_ref: str) -> dict[str, Any]:
    project = resolve_project(project_ref)
    project_id = project["project_id"]
    boq = _active_boq(project_id)
    boq_items = _active_boq_items(project_id)
    materials_boq_sources = _materials_boq_sources(project)
    invoices = sorted(
        list(budget_invoices_collection.find({"project_id": project_id})),
        key=_invoice_sort_key,
    )
    variations = list(
        budget_variations_collection.find({"project_id": project_id}).sort([("effective_date", -1)])
    )
    approved_variations = [item for item in variations if item.get("status") == "approved"]
    original_contract = sum(number(item.get("contract_amount")) for item in boq_items)
    variation_amount = sum(number(item.get("amount_delta")) for item in approved_variations)
    revised_contract = original_contract + variation_amount
    history = [item for item in invoices if item.get("status") in INVOICE_HISTORY_STATUSES]
    cumulative_certified = sum(number(item.get("certified_current_amount")) for item in history)
    cumulative_paid = sum(
        number(payment.get("amount"))
        for item in invoices
        for payment in (item.get("payments") or [])
    )
    latest = invoices[-1] if invoices else {}
    latest_comparison = latest.get("comparison") or {}
    cumulative_claimed = number(latest_comparison.get("cumulative_claimed_amount"))
    current_payable = number(latest_comparison.get("recommended_payable_amount"))
    physical_progress = number(latest_comparison.get("project_physical_progress_percentage"))
    if not invoices:
        try:
            physical_progress = number(
                (build_baseline_comparison(project_id) or {}).get("summary", {}).get("actual_percent")
            )
        except Exception:
            physical_progress = 0.0
    try:
        material_summary = get_material_summary(project_id)
    except Exception:
        material_summary = {}
    committed_value = number((material_summary.get("totals") or {}).get("committed_value"))
    exceptions: list[dict[str, Any]] = []
    for invoice in reversed(invoices):
        for result in invoice.get("verification_results") or []:
            status = str(result.get("verification_status") or "")
            if status == "verified":
                continue
            exceptions.append(
                {
                    "exception_id": f"{invoice.get('invoice_id')}:{result.get('line_id')}",
                    "invoice_id": invoice.get("invoice_id"),
                    "invoice_number": invoice.get("invoice_number"),
                    "line_id": result.get("line_id"),
                    "item_number": result.get("item_number") or result.get("boq_item_number"),
                    "description": result.get("description") or result.get("boq_description"),
                    "status": status,
                    "amount": result.get("current_claimed_amount"),
                    "variance_amount": result.get("variance_amount"),
                    "reasons": result.get("verification_reasons") or [],
                }
            )
    financial_curve: list[dict[str, Any]] = []
    certified_running = 0.0
    paid_running = 0.0
    for invoice in invoices:
        if invoice.get("status") in INVOICE_HISTORY_STATUSES:
            certified_running += number(invoice.get("certified_current_amount"))
        paid_running += sum(number(item.get("amount")) for item in invoice.get("payments") or [])
        comparison = invoice.get("comparison") or {}
        financial_curve.append(
            {
                "date": invoice.get("billing_cutoff_date") or invoice.get("invoice_date"),
                "invoice_id": invoice.get("invoice_id"),
                "claimed_percent": rounded(number(comparison.get("cumulative_claimed_amount")) / revised_contract * 100 if revised_contract else 0.0),
                "certified_percent": rounded(certified_running / revised_contract * 100 if revised_contract else 0.0),
                "paid_percent": rounded(paid_running / revised_contract * 100 if revised_contract else 0.0),
                "physical_percent": rounded(number(comparison.get("project_physical_progress_percentage"))),
            }
        )
    active_boq_public = _public(boq)
    if active_boq_public:
        active_boq_public["currency"] = (
            project_currency_code(project)
            or active_boq_public.get("currency")
            or ""
        )
        active_boq_public["revision"] = (
            active_boq_public.get("revision")
            or revision_from_filename(active_boq_public.get("original_filename"))
        )
        active_boq_public["summary"] = {
            **dict(active_boq_public.get("summary") or {}),
            "original_contract_amount": rounded(original_contract),
            "approved_variation_amount": rounded(variation_amount),
            "revised_contract_amount": rounded(revised_contract),
            "approved_variation_count": len(approved_variations),
            "pending_variation_count": sum(1 for item in variations if item.get("status") == "pending"),
        }
    return {
        "project_id": project_id,
        "site_name": project["site_name"],
        "currency": str(
            project_currency_code(project)
            or (boq or {}).get("currency")
            or latest.get("currency")
            or ""
        ),
        "summary": {
            "original_contract_amount": rounded(original_contract),
            "approved_variation_amount": rounded(variation_amount),
            "revised_contract_amount": rounded(revised_contract),
            "committed_value": rounded(committed_value),
            "cumulative_claimed_amount": rounded(cumulative_claimed),
            "cumulative_certified_amount": rounded(cumulative_certified),
            "cumulative_paid_amount": rounded(cumulative_paid),
            "current_recommended_payable": rounded(current_payable),
            "remaining_contract_value": rounded(max(revised_contract - cumulative_certified, 0.0)),
            "claimed_percentage": rounded(cumulative_claimed / revised_contract * 100 if revised_contract else 0.0),
            "certified_percentage": rounded(cumulative_certified / revised_contract * 100 if revised_contract else 0.0),
            "paid_percentage": rounded(cumulative_paid / revised_contract * 100 if revised_contract else 0.0),
            "physical_progress_percentage": rounded(physical_progress),
            "financial_physical_variance_percentage": rounded(
                cumulative_claimed / revised_contract * 100 - physical_progress
                if revised_contract
                else -physical_progress
            ),
            "retention_amount": rounded(latest_comparison.get("retention_amount")),
            "advance_recovery_amount": rounded(latest_comparison.get("advance_recovery_amount")),
            "vat_amount": rounded(latest_comparison.get("vat_amount")),
            "invoice_count": len(invoices),
            "exception_count": len(exceptions),
        },
        "active_boq": active_boq_public,
        "materials_boq_sources": materials_boq_sources,
        "boq_items": [_public(item) for item in boq_items],
        "variations": [_public(item) for item in variations],
        "invoices": [_public(item) for item in reversed(invoices)],
        "exceptions": exceptions,
        "financial_curve": financial_curve,
        "material_summary": public_value(material_summary),
    }
